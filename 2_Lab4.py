"""
Name: Kaleab Alemu and Manogya Aryal
DSC 200
Lab 4: Working with Excel Files

This program reads data from "Lab4Data.xlsx" which contains data collected from countries regarding child abuse in those
respective countries and creates a CSV file containing the country name, category of child abuse and category total.

Due Date: Oct 4, 2023
"""

# We import the openpyxl module to read the Excel file and the csv module to write the output to a csv file.
import openpyxl as op
import csv


# This is our class DSCLab4. It has 3 functions: get_categories, get_values and get_countries. It also has a
# constructor that takes in the csv file name and the worksheet as parameters.
class DSCLab4:
    def __init__(self, csvFile, ws):  # This is the constructor for our class.
        self.csvFile = csvFile  # This is the name of the csv file we will be writing to.
        self.ws = ws # This is the worksheet we will be working with.
        self.row_number = 0 #this will count the number of rows in our csv file

    # This function "get_categories", given the worksheet "Table 9" as a parameter, returns a list of categories of
    # child abuse
    def get_categories(self):
        category_names = []  # initialize our category_names list to an empty list

        # iterate through the rows and columns to extract the category names
        for cols in self.ws.iter_rows(min_row=5, max_row=7, min_col=5, max_col=31):
            row = []  # holds the list of category names for one row
            for cell in cols:  # iterate through the cells in each row
                currVal = cell.value  # holds the current value of the cell

                # iterate through the merged cells to check if the cell is in a merged cell
                for merged_cells in self.ws.merged_cells.ranges:
                    if cell.coordinate in merged_cells:  # if the cell is in a merged cell
                        currVal = merged_cells.start_cell.value  # set the current value to the value of the merged cell

                # if the cell is in row 6 and the value is in the first column, we set the current value to an empty
                # string
                if cell.row == 6 and currVal in category_names[0]:
                    currVal = ''
                row.append(currVal)
            category_names.append(row)

        categories = []
        for i in range(len(category_names[0])):  # iterate through the columns
            category = ''
            for j in range(len(category_names)):  # iterate through the rows

                # if the value is not None, we replace the newline character with an empty string
                category_names[j][i] = category_names[j][i].replace('\n', '')

                # we add the value to our category string
                category = category + '_' + category_names[j][i] if category != '' else category_names[j][i]
            categories.append(category)

        return list(dict.fromkeys(categories))  # return the list of categories without duplicates

    # This function "get_values", given the worksheet "Table 9" as a parameter, returns a list of values for each
    # country
    def get_values(self):
        values = list()  # initialize our values list to an empty list

        for row in self.ws['E15:AF211']:  # iterate through cells E15:AF211 to extract all the relevant values
            value = []  # holds the list of values for one country
            for cell in row:  # iterate through cells in each row
                # if the value is an integer or a floating point number or an en dash, we add it to our list of values
                if type(cell.value) == int or type(cell.value) == float or cell.value == chr(8211) or cell.value == chr(
                        8211) + ' ':
                    value.append(cell.value)
            values.append(value)  # append the values in one row into the values list

        return values

    # This function "get_countries", given the worksheet "Table 9" as a parameter, returns a list of countries
    def get_countries(self):
        countries = list()  # initialize our countries list to an empty list

        for row in self.ws['B15:B211']:  # iterate through cells B15:B211 to extract the country names
            for cell in row:  # iterate though the cells in each row
                countries.append(cell.value)  # add the value of each cell into our countries list
        return countries

    # This function "write_csv", given the list of countries, categories and values as parameters, writes the data into
    # a csv file
    def write_csv(self, countries, categories, values):
        csv_file = self.csvFile  # name of the csv file we will be writing to

        # open the csv file
        with open(csv_file, 'w') as csvfile:
            csvwriter = csv.writer(csvfile)  # create a csv writer object
            heading = ['CountryName', 'CategoryName', 'CategoryTotal']  # create a list of headings
            csvwriter.writerow(heading)  # write the heading into the csv file

            # iterate through the countries, categories and values and write them into the csv file
            for i in range(len(countries)):
                for j in range(len(categories)):
                    if values[i][j] != chr(8211) and values[i][j] != 0 and values[i][j] != chr(8211) + ' ':
                        csvwriter.writerow([countries[i], categories[j], values[i][j]])
                        self.row_number += 1 # increment the variable counting rows in the output csv file


def main():
    # load our workbook
    wb = op.load_workbook('data/Lab4Data.xlsx')

    # open the active worksheet, "Table 9"
    ws = wb['Table 9 ']

    # create an instance of our class DSCLab4
    lab4 = DSCLab4('aryalm1_alemuk1_lab4.csv', ws)

    # get the countries, categories and values using the functions in our class
    countries = lab4.get_countries()
    values = lab4.get_values()
    categories = lab4.get_categories()
    # write the data into a csv file
    lab4.write_csv(countries, categories, values)

    print(f"Number of rows in the output CSV file is: {lab4.row_number}") # get the number of rows in the csv file.


main()  # call the main function
